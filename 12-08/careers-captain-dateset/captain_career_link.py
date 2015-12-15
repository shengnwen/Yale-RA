__author__ = 'shengwen'
import openpyxl
import csv
import logging
import sys, traceback
logging.basicConfig( filename="./err.log",
                     filemode='w',
                     level=logging.DEBUG,
                     format= '%(asctime)s - %(levelname)s - %(message)s',
                   )
def extract_function_name():
    """Extracts failing function name from Traceback
    by Alex Martelli
    http://stackoverflow.com/questions/2380073/\
    how-to-identify-what-function-call-raise-an-exception-in-python
    """
    tb = sys.exc_info()[-1]
    stk = traceback.extract_tb(tb, 1)
    fname = stk[0][3]
    return fname

def log_exception(e):
    logging.error(
    "Function {function_name} raised {exception_class} ({exception_docstring}): {exception_message}".format(
    function_name = extract_function_name(), #this is optional
    exception_class = e.__class__,
    exception_docstring = e.__doc__,
    exception_message = e.message))

# 1604
def loadCaptainID(cap_id_dict, src = "captid_career_voy_link"):
    wb = openpyxl.load_workbook(filename=src + ".xlsx", read_only= True)
    ws = wb.get_sheet_by_name(src)
    rowN = ws.max_row
    colN = 4
    # for r_index in range(2, 100):
    for r_index in range(2, rowN + 1):
        # print "row:", r_index
        try:
            val= ws.cell(row = r_index, column = 1).value
            val = val.replace(",", " ")
            val = val.replace("  ", " ")
            id_cap_dict[ int(ws.cell(row = r_index, column = 2).value)] =  val
            val= ws.cell(row = r_index, column = 3).value
            val = val.replace(",", " ")
            val = val.replace("  ", " ")
            id_cap_dict[ int(ws.cell(row = r_index, column = 4).value)] = val
        except Exception as e:
            print "row", r_index, "has error, skipped!", val, e
            log_exception(e)
    res = "id, name\n"
    for key in id_cap_dict:
        res += str(key) + "," + id_cap_dict[key] + "\n"
    with open("id_cap.csv","w+") as fout:
        fout.writelines(res)

    # for row in ws.rows:
    #     for cell in row:
    #         print cell.value
    # 12952
def loadCapInfo(cap_info_dict, src="careers_clean"):
    wb = openpyxl.load_workbook(filename=src + ".xlsx", read_only= True)
    ws = wb.get_sheet_by_name(src)
    rowN = ws.max_row
    print rowN
    #birthloc_capt1	POCC_capt1	marriage_capt2 lastname_capt2
    colIndices = [4, 6, 9] # left for last name
    for r_index in range(3, rowN + 1):
    # for r_index in range(3, 100):
        try:
            name = ws.cell(row = r_index, column = 2).value
            name = name.replace("=", " ")
            name = name.replace("  ", " ")
            if name not in cap_info_dict:
                cap_info_dict[name] = ["", "", "", ""]
            print "row:", r_index
            cap_info_dict[name][3] = name.split()[1]
            for index in range(3):
                c_index = colIndices[index]
                val = ws.cell(row = r_index, column = c_index).value
                if val is not None:
                    cap_info_dict[name][index] = val
        except Exception as e:
            print "row", r_index, "has error, skipped!", e
            log_exception(e)
    res = "name, birth_loc, pocc, marriage, lastname\n"
    for key in cap_info_dict:
        res += str(key) + "," + ",".join(cap_info_dict[key]) + "\n"
    with open("cap_info.csv","w+") as fout:
        fout.writelines(res)
        # for c_index in colIndices:
        #     val = ws.cell(row = r_index, column = c_index).value
        #     if val is None:
        #         cap_info_dict[name].append("")
        #     else:
        #         cap_info_dict[name].append(val)

def link_cap_id_info(id_src="./id_cap.csv", info_src="./cap_info.csv"):
    with open(id_src) as fin:
        lines = fin.read().splitlines()
    id_cap = {}
    for line in lines[1:]:
        data = line.split(",")
        id_cap[data[0].strip()] = [data[1].strip().replace("  "," ")]
    # print id_cap
    cap_info = {}
    with open(info_src) as fin:
        lines = fin.read().splitlines()
    print
    for line in lines[1:]:
        data = line.split(",")
        cap_info[data[0].strip().replace("  "," ")] = [data[1].strip(), data[2].strip(), data[3].strip(), data[4].strip()]
    # print cap_info
    for data in id_cap:
        name = id_cap[data][0]
        if name in cap_info:
            id_cap[data].append(cap_info[name][0])
            id_cap[data].append(cap_info[name][1])
            id_cap[data].append(cap_info[name][2])
            id_cap[data].append(cap_info[name][3])
        else:
            print name, "not exist"
            id_cap[data].append("")
            id_cap[data].append("")
            id_cap[data].append("")
            id_cap[data].append("")
    res = "id, name, birth_loc, pocc, marriage, lastname\n"
    for key in id_cap:
        line = key+"," + ",".join(id_cap[key]) + "\n"
        res += line
    with open("id_cap_info.csv", "w+") as fout:
        fout.writelines(res)
def fillCaptainNetwork(link_table, src="captain_captain_network"):
    # wb = openpyxl.load_workbook(filename=src + ".xlsx")
    cap_id_indices = [2, 3]
    col_indices_cap1 = [i for i in range(10, 18, 2)]
    col_indices_cap2 = [i for i in range(11, 18, 2)]
    with open(src + ".csv") as fin:
        lines = fin.read().splitlines()
    res = [lines[0] + "\n"]
    print link_table["2"]
    print res
    for line in lines[1:]:
        try:
            data = line.split(",")
            capid1 = data[2]
            capid2 = data[3]
            if capid1 in link_table:
                    data[10] = link_table[capid1][1]
                    data[12] = link_table[capid1][2]
                    data[14] = link_table[capid1][3]
                    data[16] = link_table[capid1][4]
            if capid2 in link_table:
                    data[11] = link_table[capid2][1]
                    data[13] = link_table[capid2][2]
                    data[15] = link_table[capid2][3]
                    data[17] = link_table[capid2][4]
            print data
            res.append(",".join(data) + "\n")
        except:
            print "key", line, " has error, skipped!"
    with open(src + "-fill-in.csv","w+") as fout:
        fout.writelines(res)
    # print link_table["2"]
    # ws = wb.get_sheet_by_name(src)
    # rowN = ws.max_row
    # print rowN
    # for r_index in range(3, 10):
    #     capid1 = str(ws.cell(row=r_index, column=3).value)
    #     capid2 = str(ws.cell(row=r_index, column=4).value)
    #     print "captain 1:", capid1
    #     print "captain 2:", capid2
    #     #birthloc_capt1	POCC_capt1	marriage_capt2 lastname_capt2
    #     dic_index = 1
    #     if capid1 in link_table:
    #         print "enter row:", r_index, ",captain1", capid1
    #         for col_index in col_indices_cap1:
    #             print link_table[capid1][dic_index]
    #             ws.cell(row=r_index, column=col_index).value = link_table[capid1][dic_index]
    #             dic_index += 1
    #     dic_index = 1
    #     if capid2 in link_table:
    #         print "enter row:", r_index, ",captain2", capid2
    #         for col_index in col_indices_cap2:
    #             ws.cell(row=r_index, column=col_index).value = link_table[capid2][dic_index]
    #             dic_index += 1
    # wb.save(src + ".xlsx")
def fillCaptainNetworkCSV(src="captain_captain_network"):
    # wb = openpyxl.load_workbook(filename=src + ".xlsx")
    link_table = {}
    with open("./id_cap_info.csv") as fin:
        lines = fin.read().splitlines()
    for line in lines[1:]:
        data = line.split(",")
        link_table[data[0]] = [i for i in data[1:]]
    cap_id_indices = [2, 3]
    col_indices_cap1 = [i for i in range(10, 18, 2)]
    col_indices_cap2 = [i for i in range(11, 18, 2)]
    with open(src + ".csv") as fin:
        lines = fin.read().splitlines()
    res = [lines[0] + "\n"]
    print link_table["2"]
    print res
    for line in lines[1:]:
        try:
            data = line.split(",")
            capid1 = data[2]
            capid2 = data[3]
            if capid1 in link_table:
                    data[10] = link_table[capid1][1]
                    data[12] = link_table[capid1][2]
                    data[14] = link_table[capid1][3]
                    data[16] = link_table[capid1][4]
            if capid2 in link_table:
                    data[11] = link_table[capid2][1]
                    data[13] = link_table[capid2][2]
                    data[15] = link_table[capid2][3]
                    data[17] = link_table[capid2][4]
            print data
            res.append(",".join(data) + "\n")
        except:
            print "key", line, " has error, skipped!"
    with open(src + "-fill-in.csv","w+") as fout:
        fout.writelines(res)
    # print link_table["2"]
    # ws = wb.get_sheet_by_name(src)
    # rowN = ws.max_row
    # print rowN
    # for r_index in range(3, 10):
    #     capid1 = str(ws.cell(row=r_index, column=3).value)
    #     capid2 = str(ws.cell(row=r_index, column=4).value)
    #     print "captain 1:", capid1
    #     print "captain 2:", capid2
    #     #birthloc_capt1	POCC_capt1	marriage_capt2 lastname_capt2
    #     dic_index = 1
    #     if capid1 in link_table:
    #         print "enter row:", r_index, ",captain1", capid1
    #         for col_index in col_indices_cap1:
    #             print link_table[capid1][dic_index]
    #             ws.cell(row=r_index, column=col_index).value = link_table[capid1][dic_index]
    #             dic_index += 1
    #     dic_index = 1
    #     if capid2 in link_table:
    #         print "enter row:", r_index, ",captain2", capid2
    #         for col_index in col_indices_cap2:
    #             ws.cell(row=r_index, column=col_index).value = link_table[capid2][dic_index]
    #             dic_index += 1
    # wb.save(src + ".xlsx")
def store_table(id_cap_dict, cap_info_dict):
    id_cap_src = "./id_cap_table"
    cap_info_src = "./cap_info_dict"
    res = []
    for key in id_cap_dict:
        try:
            name = id_cap_dict[key]
            if name in cap_info_dict:
                res.append(",".join([str(key)] + [name] + cap_info_dict[name]) + "\n")
        except:
            print "key", key, " has error, skipped!"
    with open("link_table.csv","w+") as fout:
        fout.writelines(res)
def get_link_table(link_table):
    with open("link_table.csv") as fin:
        lines = fin.read().splitlines()
    for line in lines:
        try:
            data = line.split(",")
            link_table[data[0]] = [i for i in data[1:]]
        except:
            print "key", line, " has error, skipped!"
def enter_into_network():
    pass
id_cap_dict = {}
cap_info_dict = {}
# link_table = {}
# loadCaptainID(id_cap_dict)
# loadCapInfo(id_cap_dict, cap_info_dict)
# loadCaptainID(id_cap_dict)
# print "loadCaptainID done"
#loadCapInfo(cap_info_dict)
# print "loadCaptainInfo Done"
# store_table(id_cap_dict, cap_info_dict)
# print "store_link_table"
# get_table(link_table)
# print "link_table done"
# print link_table
# print cap_info_dict
# get_link_table(link_table)
#
# fillCaptainNetwork(link_table)
# fillCaptainNetwork(id_cap_dict, cap_info_dict)

#link_cap_id_info()
fillCaptainNetworkCSV()

